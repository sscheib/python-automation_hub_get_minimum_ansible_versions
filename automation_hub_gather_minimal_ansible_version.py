#!/usr/bin/env python3
# -*- coding: UTF-8 -*-

'''
  Description:
    This script iterates through all Ansible Collections which are available from the Ansible Automation 
    Hub (console.redhat.com) in both the validated content and the certified content and prints out the
    minimal required Ansible Core version for the latest available collection version.

    Optionally, the data can be written to a spreadsheet (.xlsx).
'''

import requests
import logging
import sys
import os
import yaml
import re
import openpyxl
import errno
from enum import Enum
from datetime import datetime
from datetime import timedelta
from pathlib import Path
from argparse import ArgumentParser
from getpass import getpass
from pprint import pformat, pprint
from packaging.version import Version, parse

__author__ = 'Steffen Scheib'
__copyright__ = 'Copyright 2025, Steffen Scheib'
__credits__ = ['Steffen Scheib']
__license__ = 'GPLv2 or later'
__version__ = '0.2.0'
__maintainer__ = 'Steffen Scheib'
__email__ = 'steffen@scheib.me'
__status__ = 'Development'

LOG = logging.getLogger(os.path.basename(os.path.splitext(__file__)[0]))
API_URL = 'https://console.redhat.com'
API_USERNAME = ''
API_PASSWORD = ''

class HttpRequestType(Enum):
    '''Representation of the different HTTP request types'''
    GET = 1
    POST = 2
    PUT = 3
    DELETE = 4


def query_api(http_request_type: HttpRequestType, location: str, data: str = None) -> dict:
  """Queries the API
  Queries the Automation Hub API and returns the result as JSON formatted string.
  HTTP types supported are: GET, POST, PUT, DELETE.

  Args:
      http_request_type (str): The HTTP request type to use. Supported are GET, POST, PUT, DELETE
      location (str): Location to query (Example: content_views/1)
      data (str, optional): The optional payload to deliver with the HTTP request

  Returns:
      dict: The resulting response from the HTTP requests as JSON formatted string (=dict)

  Raises:
      ValueError: If the first argument is None or not given
      TypeError: If the first argument is not an instance of HttpRequestType
      ValueError: If the second argument is None or not given
      TypeError: If the second argument is not an instance of ApiType
      TypeError: If the optional third argument is given, but is not a string
      ValueError: If the optional third argument is given, but is not a JSON formatted string
      ValueError: If the given HTTP request type is not supported
      HTTPError: If the request returns with an unsuccessful status code
      ConnectionError: If a connection to the  API cannot be established (DNS failure, connection
                       refused, etc)
      Timeout: If the request exceeds the maximum time in which it didn't receive any data
      RequestException: If the HTTP request fails for another reason
      RuntimeError: If the HTTP request fails for some reason
  """
  # check existence and type of the first argument
  if not http_request_type or http_request_type is None:
      raise ValueError(f'Given value for the first argument (\'http_request_type\') is empty (or None).')
  elif not isinstance(http_request_type, HttpRequestType):
      raise TypeError(f'Given value for the first argument (\'http_request_type\') is not an instance '
                      f'of HttpRequestType. Type of value is {type(http_request_type)}.')

  # check existence and type of the second argument
  if not location or location is None:
      raise ValueError(f'Given value for the second argument (\'location\') is empty (or None).')
  elif not isinstance(location, str):
      raise TypeError(f'Given value for the second argument (\'location\') is not a string. Type of value '
                      f'is {type(location)}.')

  if data is not None:
    LOG.debug(f'Using HTTP {http_request_type.name} on {API_URL + location} payload {pformat(data)}')
  else:
    LOG.debug(f'Using HTTP {http_request_type.name} on {API_URL + location}')


  # do the HTTP request
  auth = (API_USERNAME, API_PASSWORD)
  try:
      if http_request_type is HttpRequestType.GET:
          response = requests.get(API_URL + location,
                                  data=data,
                                  auth=auth,
                                  headers={'content-type': 'application/json'})
      elif http_request_type is HttpRequestType.POST:
          response = requests.post(API_URL + location,
                                   data=data,
                                   auth=auth,
                                   headers={'content-type': 'application/json'})
      elif http_request_type is HttpRequestType.PUT:
          response = requests.put(API_URL + location,
                                  data=data,
                                  auth=auth,
                                  headers={'content-type': 'application/json'})
      elif http_request_type is HttpRequestType.DELETE:
          response = requests.delete(API_URL + location,
                                     data=data,
                                     auth=auth,
                                     headers={'content-type': 'application/json'})
      else:
          raise ValueError(f'Given HTTP request type is not supported! Given is {http_request_type.name}.')

      return response.json()
      response.raise_for_status()
  except requests.exceptions.HTTPError as http_error:
      raise requests.exceptions.HTTPError(f'The HTTP {http_request_type.name} request failed with an HTTPError. '
                                          f'Following the complete error:'
                                          f' {http_error}')
  except requests.exceptions.ConnectionError as connection_error:
      raise requests.exceptions.ConnectionError(f'Unable to connect to the configured API {API_URL}. '
                                                f'Following the complete error: '
                                                f'{connection_error}')
  except requests.exceptions.ReadTimeout as read_timeout_error:
      raise requests.exceptions.ReadTimeout(f'The HTTP {http_request_type.name} request timed out.' 
                                            f'Following the complete error: {read_timeout_error}')
  except requests.exceptions.Timeout as timeout_error:
      raise requests.exceptions.Timeout(f'Timeout of the HTTP {http_request_type.name} request has been reached. '
                                        f'Following the complete error: '
                                        f'{timeout_error}')
  except requests.exceptions.RequestException as request_exception:
      raise requests.exceptions.RequestException(f'The HTTP {http_request_type.name} request failed. Following '
                                                 f'the complete error: '
                                                 f'{request_exception}')

  if not response.ok:
      raise RuntimeError(f'Last {http_request_type.name} request failed. Request returned with '
                         f'HTTP code {response.status_code}')

  # return the response as JSON
  return response.json()

parser = ArgumentParser()
parser.add_argument('--api-url', dest='api_url',
                    help='The base URL of the API',
                    default='https://console.redhat.com', required=False)
parser.add_argument('--api-username', dest='api_username',
                    help='Username to authenticate against the API',
                    required=True)
parser.add_argument('--api-password', dest='api_password',
                    help='Password for the user to authenticate against the API',
                    required=False)
parser.add_argument('--workbook-path', dest='workbook_path', default='/tmp/collections.xlsx',
                    help='Path to store the workbook (xlsx)',
                    required=False)
parser.add_argument('--no-workbook', action='store_false', dest='write_workbook', default=True,
                    help='Whether to write a workbook',
                    required=False)
parser.add_argument('--clear-workbook', action='store_true', dest='clear_workbook', default=True,
                    help='Remove the workbook file on start',
                    required=False)
parser.add_argument('--ignore-authors', action='store_true', dest='ignore_authors', default=False,
                    help='Add authors of the collection to the workbook',
                    required=False)
args = parser.parse_args()

# set the log level
LOG.setLevel(logging.INFO)

# create console handler
console_handler = logging.StreamHandler()
console_handler.setLevel(logging.INFO)
console_formatter = logging.Formatter('[%(asctime)s] %(name)-12s %(levelname)-8s: %(funcName)-50s: %(message)s')
console_handler.setFormatter(console_formatter)
LOG.addHandler(console_handler)

API_URL = args.api_url
API_USERNAME = args.api_username
if args.api_password:
    API_PASSWORD = args.api_password
else:
    API_PASSWORD = getpass(f'Password for user {API_USERNAME}: ')

# set the initial href for each repository type
hrefs = {
  'validated': '/api/automation-hub/v3/plugin/ansible/content/validated/collections/index/?limit=100',
  'certified': '/api/automation-hub/v3/plugin/ansible/content/published/collections/index/?limit=100'
}

if args.write_workbook:
    workbook_path = args.workbook_path

    if args.clear_workbook:
        try:
            os.remove(workbook_path)
        except OSError as oserror_exception:
            if oserror_exception.errno != errno.ENOENT: # errno.ENOENT = no such file or directory
                raise

    workbook = openpyxl.Workbook()
    worksheet = workbook.create_sheet('Collections Ansible versions')
    # create the header column for the worksheet
    worksheet.cell(column=1, row=1, value='Collection Name')
    worksheet.cell(column=2, row=1, value='Repository')
    worksheet.cell(column=3, row=1, value='Ansible Version')
    worksheet.cell(column=4, row=1, value='Downloads')

    if not args.ignore_authors:
        worksheet.cell(column=5, row=1, value='Authors')

ansible_versions = dict()

# initial row accomodates for the header column
row = 2

# iterate over all repositories
for collection_repo, initial_href in hrefs.items():

    href = initial_href
    while True:
        result = query_api(HttpRequestType.GET, href)

        # iterate over each collection
        for collection in result['data']:

            collection_name = collection['name']
            collection_namespace = collection['namespace']
            collection_fqcn = f'{collection_namespace}.{collection_name}'
            download_count = collection['download_count']
            highest_version = query_api(HttpRequestType.GET, collection['highest_version']['href'])
            ansible_version = Version(re.sub(',[0-9.]+', '', re.sub('>=|<=|>|<|,[0-9.]+', '', highest_version['requires_ansible'])))
            authors = ', '.join(highest_version['metadata']['authors'])
            ansible_minor_version = f'{ansible_version.major}.{ansible_version.minor}'

            if ansible_minor_version in ansible_versions:
                ansible_versions[ansible_minor_version]['collections'].append(
                    {
                        'name': collection_fqcn,
                        'download_count': download_count
                    }
                )
            else:
                ansible_versions[ansible_minor_version] = {
                    'collections': [
                        {
                            'name': collection_fqcn,
                            'download_count': download_count
                        }
                    ]
                }

            if args.write_workbook:            
                worksheet.cell(column=1, row=row, value=collection_fqcn)
                worksheet.cell(column=2, row=row, value=collection_repo)
                worksheet.cell(column=3, row=row, value=ansible_minor_version)
                worksheet.cell(column=4, row=row, value=str(download_count))

                if not args.ignore_authors:
                    worksheet.cell(column=5, row=row, value=str(authors))

            pprint(f"collection {collection_fqcn} (#{download_count} downloads): -> {str(ansible_minor_version)}")
            row = row + 1

        # we are done once no next link is given
        if result['links']['next'] is None:
            break

        # assign new href
        href = result['links']['next']

if args.write_workbook:
    # delete the initially created sheet
    workbook.remove(workbook['Sheet'])

    # save the workbook
    workbook.save(workbook_path)

sys.exit(0)
